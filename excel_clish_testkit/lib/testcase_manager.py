import os
import copy

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

import config
import constants
from exceptions import InvalidExcelException


class TestcaseManager:
    def __init__(self, tcname):
        self.tcname = tcname
        self.excel_path = os.path.join(
            config.TEST_SRC_DIR, f'{self.tcname}.xlsx'
        )
        self.wb_obj = None
        self.ws_obj = None
        self.cmd_rows = []
        self.headers = []

    def get_row(self, row_obj):
        return [cell.value for cell in row_obj][:len(self.headers)]

    def get_headers(self):
        return [cell.value.strip() for cell in self.ws_obj[1] \
                if cell.value not in constants.CELL_EMPTY_VALUES]

    def save_headers(self):
        for (c, h) in enumerate(self.headers):
            cell = self.ws_obj.cell(row=1, column=c + 1)
            cell.value = h
            cell.font = Font(color='FFFFFF')
            background_color = 'C0504D'
            cell.fill = PatternFill(
                fill_type="solid",
                start_color=background_color,
                end_color=background_color
            )

    def set_new_headers(self):
        self.headers = copy.copy(constants.TEST_HEADER)
        self.save_headers()

    def check_is_valid_excel(self):
        if constants.SUCCESS_COL not in self.headers \
                or constants.ERROR_COL not in self.headers \
                or constants.TEST_RESULT_COL not in self.headers:
            raise InvalidExcelException(
                f"The {os.path.basename(self.excel_path)} Excel needs {constants.SUCCESS_COL}, " \
                + f"{constants.ERROR_COL} and {constants.TEST_RESULT_COL}"
            )

    def load_tc(self):
        print(f"Loading {self.tcname}")
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f'{self.excel_path} file not found')

        self.wb_obj = load_workbook(self.excel_path)
        self.ws_obj = self.wb_obj.active

        self.headers = self.get_headers()

        self.check_is_valid_excel()

    def get_tests(self):
        print(f'{self.tcname} | Loading tests')
        self.load_tc()

        for row_obj in self.ws_obj.iter_rows(min_row=2):
            row_cells = self.get_row(row_obj)

            if not row_cells or not row_cells[0]:
                print(f'Empty row found in {self.excel_path}')
                break

            self.cmd_rows.append(row_cells)

        return self.cmd_rows

    def save_excel(self):
        try:
            self.wb_obj.save(self.excel_path)
        except PermissionError as exc:
            print(
                f"{self.tcname} | Failed to update {self.excel_path} excel. "
                + "Please make sure the file is not open"
            )

    def update_rows(self, test_rows):
        print(f"{self.tcname} | Updating rows")

        for (r, test_row) in enumerate(test_rows):
            for (c, cell) in enumerate(test_row):
                self.ws_obj.cell(row=r + 1 + constants.EXCEL_ROW_OFFSET, column=c + 1).value = cell

    # New tc generation
    def create_new_tc(self):
        self.wb_obj = Workbook()
        self.ws_obj = self.wb_obj.active

        self.set_new_headers()
        self.save_excel()

    def append_tc_headers(self):
        print('Setting required headers')

        if constants.SUCCESS_COL not in self.headers:
            self.headers.append(constants.SUCCESS_COL)
        if constants.ERROR_COL not in self.headers:
            self.headers.append(constants.ERROR_COL)
        if constants.TEST_RESULT_COL not in self.headers:
            self.headers.append(constants.TEST_RESULT_COL)
        if constants.SKIP_COL not in self.headers:
            self.headers.append(constants.SKIP_COL)
        self.save_headers()

    def prepare_tc(self):
        print(f"Preparing {self.tcname}")
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f'{self.excel_path} file not found')

        self.wb_obj = load_workbook(self.excel_path)
        self.ws_obj = self.wb_obj.active

        self.headers = self.get_headers()
        self.append_tc_headers()
        self.save_excel()

    @staticmethod
    def prepare_tcs():
        print(f'Preparing testcases in {config.TEST_SRC_DIR} location')

        for entry in os.scandir(config.TEST_SRC_DIR):
            if entry.is_file() and entry.name.endswith('.xlsx') and not entry.name.startswith('~$'):
                tcname = entry.name.split('.')[0]
                tc_mgr = TestcaseManager(tcname)
                tc_mgr.prepare_tc()


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "action", help="Create/Prepare a TC Excel for Testing",
        choices=['new', 'prepare']
    )
    parser.add_argument(
        "tcname", help="Testcase name of TC Excel"
    )

    args = parser.parse_args()

    tcmgr = TestcaseManager(args.tcname)

    if args.action == 'new':
        tcmgr.create_new_tc()
    else:
        try:
            tcmgr.prepare_tc()
        except FileNotFoundError:
            print(args.tcname, "TC Excel Not Found")
