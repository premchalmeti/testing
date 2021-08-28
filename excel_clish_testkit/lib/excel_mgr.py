import os

from openpyxl import load_workbook

import utils
import config
import constants
from exceptions import InvalidExcelException


class ExcelManager:
    """
    Parse the `config.CURRENT_TC` excel file and is responsible for following things,
    - read currently running TC (`constants.CURRENT_TC`)
    - return test_rows to test_runner instance
    - executes tests
    - update test rows with results and update TC excel

    >>> parser = ExcelManager("device.xlsx")
    >>> test_results = parser.get_tests()
    >>> parser.update_test_result(test_results)
    """
    # CMD_CELLS: columns used for cmd
    CMD_CELLS = constants.EXCEL_CMD_CELL_LIMIT

    # ROW_OFFSET: start of commands
    ROW_OFFSET = constants.EXCEL_ROW_OFFSET

    def __init__(self):
        self.logger = utils.get_logger(f"{constants.CURRENT_TC}.{self.__class__.__name__}")
        self.excel_path = os.path.join(
            config.TEST_SRC_DIR, f'{constants.CURRENT_TC}.xlsx'
        )
        self.logger.debug(f'Using excel {self.excel_path} testcase path')
        self.logger.debug(f"{self.CMD_CELLS} columns will be used for CMD")
        self.wb_obj = None
        self.ws_obj = None
        self.cmd_rows = []
        self.headers = []

    @staticmethod
    def get_cmd(cmd_cells):
        plain_cmd_cells = []

        for c in cmd_cells:
            if c in constants.CELL_EMPTY_VALUES or utils.is_param(c):
                break
            plain_cmd_cells.append(c)

        return plain_cmd_cells

    @staticmethod
    def get_complete_cmd(row_cells):
        cmd_cells = []
        for c in row_cells[:ExcelManager.CMD_CELLS+1]:
            if c in constants.CELL_EMPTY_VALUES:
                break
            cmd_cells.append(c)
        return cmd_cells

    def get_row(self, row_obj):
        return [cell.value for cell in row_obj][:len(self.headers)]

    def get_headers(self):
        return [cell.value.strip() for cell in self.ws_obj[1] \
                if cell.value not in constants.CELL_EMPTY_VALUES]

    def check_is_valid_excel(self):
        if constants.SUCCESS_COL not in self.headers \
                or constants.ERROR_COL not in self.headers \
                or constants.TEST_RESULT_COL not in self.headers:
            raise InvalidExcelException(
                f"The {os.path.basename(self.excel_path)} Excel needs {constants.SUCCESS_COL}, " \
                + f"{constants.ERROR_COL} and {constants.TEST_RESULT_COL}"
            )

    def get_tests(self):
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f'{self.excel_path} file not found')

        self.logger.debug("Loading excel test case")

        self.wb_obj = load_workbook(self.excel_path)
        self.ws_obj = self.wb_obj.active

        self.headers = self.get_headers()
        self.logger.debug(f"Excel headers: {self.headers}")

        self.check_is_valid_excel()

        for row_obj in self.ws_obj.iter_rows(min_row=2):
            row_cells = self.get_row(row_obj)

            self.logger.debug(f"Found test: {row_cells}")

            if not row_cells or not row_cells[0]:
                self.logger.warning(f'Empty row found in {self.excel_path}')
                break

            self.cmd_rows.append(row_cells)

        return self.cmd_rows

    def save_excel(self):
        try:
            self.wb_obj.save(self.excel_path)

            self.logger.info(f"{constants.CURRENT_TC} | Test Excel Updated!")
        except PermissionError as exc:
            self.logger.error(
                f"{constants.CURRENT_TC} | Failed to update {self.excel_path} excel. "
                + "Please make sure the file is not open",
                exc_info=True
            )

    def update_test_result(self, test_rows):
        self.logger.info(f"Updating test results")

        for (r, test_row) in enumerate(test_rows):
            for (c, cell) in enumerate(test_row):
                self.ws_obj.cell(row=r + 1 + self.ROW_OFFSET, column=c + 1).value = cell
